import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


async def export_to_excel(data, headings, filepath):

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Data"

    # =========================
    # STYLE
    # =========================

    header_font = Font(bold=True, color="FFFFFF")

    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    center_align = Alignment(horizontal="center", vertical="center")

    # =========================
    # HEADER
    # =========================

    for colno, heading in enumerate(headings, start=1):

        cell = sheet.cell(row=1, column=colno)

        cell.value = heading
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    # =========================
    # DATA
    # =========================

    for rowno, row in enumerate(data, start=2):

        for colno, cell_value in enumerate(row, start=1):

            cell = sheet.cell(row=rowno, column=colno)

            cell.value = cell_value
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # =========================
    # AUTO COLUMN WIDTH
    # =========================

    for column_cells in sheet.columns:

        length = 0
        column = column_cells[0].column_letter

        for cell in column_cells:
            try:
                if cell.value:
                    length = max(length, len(str(cell.value)))
            except:
                pass

        sheet.column_dimensions[column].width = length + 3

    # =========================
    # FREEZE HEADER
    # =========================

    sheet.freeze_panes = "A2"

    # =========================
    # FILTER
    # =========================

    sheet.auto_filter.ref = sheet.dimensions

    wb.save(filepath)